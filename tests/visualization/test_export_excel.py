from pathlib import Path

import openpyxl

from src.visualization.export_excel import build_sprint3_exports


def test_build_sprint3_exports_generates_excel_and_markdown_outputs(tmp_path: Path) -> None:
    output_dir = tmp_path / "output"
    reports_dir = tmp_path / "reports"

    result = build_sprint3_exports(
        db_path=Path("database/n100.db"),
        output_dir=output_dir,
        reports_dir=reports_dir,
    )

    assert result["screener_workbook"] == output_dir / "screener_output.xlsx"
    assert result["peer_workbook"] == output_dir / "peer_comparison.xlsx"
    assert (output_dir / "screener_output.xlsx").exists()
    assert (output_dir / "peer_comparison.xlsx").exists()
    assert (reports_dir / "final_validation_report.md").exists()
    assert (reports_dir / "performance_summary.md").exists()
    assert (reports_dir / "sprint3_summary.md").exists()
    assert (reports_dir / "architecture.md").exists()
    assert (reports_dir / "deployment.md").exists()

    workbook = openpyxl.load_workbook(output_dir / "screener_output.xlsx")
    assert "Quality Compounder" in workbook.sheetnames
    assert "Value Pick" in workbook.sheetnames
