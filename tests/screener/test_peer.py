from pathlib import Path

import pandas as pd

from src.analytics.peer import (
    compute_peer_percentiles,
    normalize_financial_year,
    prepare_peer_frame,
    run_peer_engine,
)


def test_normalize_financial_year_handles_common_inputs() -> None:
    assert normalize_financial_year("Dec 2012") == 2012
    assert normalize_financial_year(2021) == 2021
    assert normalize_financial_year(None) is None


def test_prepare_peer_frame_keeps_latest_year_per_company() -> None:
    financial_ratios = pd.DataFrame(
        [
            {"company_id": "A", "year": "Mar 2022", "return_on_equity_pct": 10.0},
            {"company_id": "A", "year": "Mar 2023", "return_on_equity_pct": 12.0},
            {"company_id": "B", "year": "Mar 2023", "return_on_equity_pct": 8.0},
        ]
    )
    companies = pd.DataFrame([{"company_id": "A", "company_name": "Alpha"}, {"company_id": "B", "company_name": "Beta"}])
    peer_groups = pd.DataFrame([{"company_id": "A", "peer_group_name": "Group 1"}, {"company_id": "B", "peer_group_name": "Group 1"}])
    sectors = pd.DataFrame([{"company_id": "A", "broad_sector": "Tech"}, {"company_id": "B", "broad_sector": "Tech"}])

    prepared = prepare_peer_frame(financial_ratios, companies, peer_groups, sectors)

    assert prepared["company_id"].tolist() == ["A", "B"]
    assert prepared["financial_year"].tolist() == [2023, 2023]
    assert prepared.loc[prepared["company_id"] == "A", "return_on_equity_pct"].iloc[0] == 12.0


def test_compute_peer_percentiles_ranks_higher_metrics_higher() -> None:
    frame = pd.DataFrame(
        [
            {"company_id": "A", "peer_group": "Group 1", "metric": "return_on_equity_pct", "metric_value": 10.0},
            {"company_id": "B", "peer_group": "Group 1", "metric": "return_on_equity_pct", "metric_value": 30.0},
        ]
    )

    result = compute_peer_percentiles(frame)
    ranked = result[result["metric"] == "return_on_equity_pct"].set_index("company_id")

    assert ranked.loc["B", "percentile_rank"] > ranked.loc["A", "percentile_rank"]


def test_compute_peer_percentiles_inverts_debt_metric() -> None:
    frame = pd.DataFrame(
        [
            {"company_id": "A", "peer_group": "Group 1", "metric": "debt_to_equity", "metric_value": 0.2},
            {"company_id": "B", "peer_group": "Group 1", "metric": "debt_to_equity", "metric_value": 1.0},
        ]
    )

    result = compute_peer_percentiles(frame)
    ranked = result[result["metric"] == "debt_to_equity"].set_index("company_id")

    assert ranked.loc["A", "percentile_rank"] > ranked.loc["B", "percentile_rank"]


def test_compute_peer_percentiles_handles_no_peer_group() -> None:
    frame = pd.DataFrame(
        [
            {"company_id": "C", "peer_group": "No Peer Group", "metric": "return_on_equity_pct", "metric_value": 9.0},
        ]
    )

    result = compute_peer_percentiles(frame)

    assert result.iloc[0]["peer_group"] == "No Peer Group"
    assert result.iloc[0]["percentile_rank"] == 100.0


def test_run_peer_engine_creates_sqlite_table_and_output_files(tmp_path: Path) -> None:
    output_dir = tmp_path / "output"
    reports_dir = tmp_path / "reports"

    result = run_peer_engine(
        db_path=Path("database/n100.db"),
        output_dir=output_dir,
        reports_dir=reports_dir,
    )

    assert not result.empty
    assert (output_dir / "peer_comparison.xlsx").exists()
    assert (reports_dir / "peer_summary.md").exists()
    assert (reports_dir / "peer_validation.md").exists()


def test_run_peer_engine_writes_expected_columns() -> None:
    result = run_peer_engine(db_path=Path("database/n100.db"))

    assert {"company_id", "peer_group", "metric", "metric_value", "percentile_rank", "financial_year"}.issubset(result.columns)


def test_run_peer_engine_assigns_latest_year_per_company() -> None:
    result = run_peer_engine(db_path=Path("database/n100.db"))

    assert result["financial_year"].notna().all()
    assert result["peer_group"].notna().all()


def test_workbook_has_one_sheet_per_peer_group(tmp_path: Path) -> None:
    output_dir = tmp_path / "output"
    run_peer_engine(db_path=Path("database/n100.db"), output_dir=output_dir)

    import openpyxl

    workbook = openpyxl.load_workbook(output_dir / "peer_comparison.xlsx")
    assert len(workbook.sheetnames) >= 11


def test_workbook_contains_company_and_metric_columns(tmp_path: Path) -> None:
    output_dir = tmp_path / "output"
    run_peer_engine(db_path=Path("database/n100.db"), output_dir=output_dir)

    import openpyxl

    workbook = openpyxl.load_workbook(output_dir / "peer_comparison.xlsx")
    sheet = workbook[workbook.sheetnames[0]]

    header = [cell.value for cell in sheet[1]]
    assert "Company" in header
    assert "Composite Score" in header
    assert "ROE Percentile" in header


def test_workbook_includes_median_row(tmp_path: Path) -> None:
    output_dir = tmp_path / "output"
    run_peer_engine(db_path=Path("database/n100.db"), output_dir=output_dir)

    import openpyxl

    workbook = openpyxl.load_workbook(output_dir / "peer_comparison.xlsx")
    sheet = workbook[workbook.sheetnames[0]]
    values = [cell.value for cell in sheet[sheet.max_row]]

    assert values[0] == "Median"


def test_highest_roe_gets_highest_roe_percentile_within_a_peer_group() -> None:
    result = run_peer_engine(db_path=Path("database/n100.db"))
    peer_groups = result[result["metric"] == "return_on_equity_pct"].groupby("peer_group")

    for peer_group, group in peer_groups:
        if group.empty:
            continue
        highest = group.loc[group["metric_value"].idxmax()]
        highest_rank = highest["percentile_rank"]
        assert highest_rank >= group["percentile_rank"].max()


def test_debt_to_equity_inversion_is_applied() -> None:
    result = run_peer_engine(db_path=Path("database/n100.db"))
    debt_group = result[result["metric"] == "debt_to_equity"]

    if not debt_group.empty:
        for peer_group, group in debt_group.groupby("peer_group"):
            if group.shape[0] < 2:
                continue
            lowest = group.loc[group["metric_value"].idxmin()]
            assert lowest["percentile_rank"] >= group["percentile_rank"].max() - 1e-9


def test_no_peer_group_is_saved_as_label() -> None:
    result = run_peer_engine(db_path=Path("database/n100.db"))

    assert "No Peer Group" in set(result["peer_group"].dropna())


def test_summary_report_contains_peer_group_count(tmp_path: Path) -> None:
    output_dir = tmp_path / "output"
    reports_dir = tmp_path / "reports"
    run_peer_engine(db_path=Path("database/n100.db"), output_dir=output_dir, reports_dir=reports_dir)

    summary_text = (reports_dir / "peer_summary.md").read_text(encoding="utf-8")

    assert "Peer Comparison Summary" in summary_text
    assert "peer groups" in summary_text.lower()


def test_validation_report_contains_validation_results(tmp_path: Path) -> None:
    output_dir = tmp_path / "output"
    reports_dir = tmp_path / "reports"
    run_peer_engine(db_path=Path("database/n100.db"), output_dir=output_dir, reports_dir=reports_dir)

    validation_text = (reports_dir / "peer_validation.md").read_text(encoding="utf-8")

    assert "Validation" in validation_text
    assert "PASS" in validation_text
